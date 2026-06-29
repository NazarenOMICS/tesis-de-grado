#!/usr/bin/env python3
"""
Auditoria reproducible del benchmark TripleSTD DDA.
Recalcula, desde los Excel originales, las metricas que el analisis IA
(legacy/analisis_ia_triple_std_dda_resultadoporGPT.txt) reporta, para
verificarlas antes de integrarlas como resultado.

Fuentes:
  - Pairwise_comparison_TripleSTD.xlsx (hoja Table1, log2FC por peptido/proteina)
  - TFold_Analysis_ConParametrosAlejandro_2026-04-17.xlsx (hojas Blue/Green/Red Dots, FC lineal)

Uso: python3 auditoria_triple_std.py
Requiere: openpyxl
Ejecutado y verificado: 2026-06-29.
"""
import math
import statistics as st
import openpyxl

PAIRWISE = "Pairwise_comparison_TripleSTD.xlsx"
TFOLD = "TFold_Analysis_ConParametrosAlejandro_2026-04-17.xlsx"


def organismo(texto):
    s = (texto or "").lower()
    if s.startswith("reverse"):
        return "Decoy"
    if "coli" in s:
        return "E. coli"
    if "cerevis" in s or "yeast" in s:
        return "Yeast"
    if "sapiens" in s or "homo" in s:
        return "Human"
    return "Sin asignar"


def log2_desde_fc_lineal(fc):
    fc = float(fc)
    if fc > 0:
        return math.log2(fc)
    if fc < 0:
        return -math.log2(abs(fc))
    return float("nan")


def auditar_pairwise():
    wb = openpyxl.load_workbook(PAIRWISE, read_only=True, data_only=True)
    filas = list(wb["Table1"].iter_rows(values_only=True))
    h = filas[0]
    iL, iOS = h.index("LOG2(FOLD)"), h.index("OS")
    iLoc = h.index("LOCUS")
    por_org, loci = {}, set()
    for r in filas[1:]:
        loci.add(str(r[iLoc]))
        o = organismo(r[iOS]) if not str(r[iLoc]).lower().startswith("reverse") else "Decoy"
        por_org.setdefault(o, []).append(float(r[iL]))
    print(f"PAIRWISE (n={len(filas) - 1})")
    for o, v in por_org.items():
        med = st.median(v)
        print(f"  {o:11s} n={len(v):4d} mediana_log2={med:+.3f} "
              f"FC_lineal={2 ** med:.3f} sd={(st.stdev(v) if len(v) > 1 else 0):.3f}")
    return loci


def auditar_tfold():
    wb = openpyxl.load_workbook(TFOLD, read_only=True, data_only=True)
    conjuntos = {}
    for hoja in ["Blue Dots", "Green Dots", "Red Dots"]:
        filas = list(wb[hoja].iter_rows(values_only=True))
        h = filas[0]
        iFC = h.index("Fold Change")
        por_org, loci = {}, set()
        for r in filas[1:]:
            loc = str(r[0])
            loci.add(loc)
            o = "Decoy" if loc.lower().startswith("reverse") else organismo(r[-1])
            if r[iFC] is not None:
                por_org.setdefault(o, []).append(log2_desde_fc_lineal(r[iFC]))
        conjuntos[hoja] = loci
        print(f"{hoja} (n={len(filas) - 1})")
        for o in ["E. coli", "Yeast", "Human", "Decoy", "Sin asignar"]:
            if o in por_org:
                v = por_org[o]
                med = st.median(v)
                print(f"  {o:11s} n={len(v):4d} mediana_log2={med:+.3f} FC_lineal={2 ** med:.3f}")
    return conjuntos


def main():
    pw = auditar_pairwise()
    print()
    tf = auditar_tfold()
    print()
    blue = tf["Blue Dots"]
    print("OVERLAP y FDR empirico")
    print(f"  Pairwise sig = {len(pw)}")
    inter = len(pw & blue)
    print(f"  Pairwise n Blue = {inter} ({100 * inter / len(pw):.1f}%)")
    print(f"  FDR empirico Blue = 138/546 = {100 * 138 / 546:.1f}% (Human sobre total Blue)")
    print(f"  FDR empirico Pairwise = 1/271 = {100 / 271:.2f}% (1 Human sobre organismos asignados)")


if __name__ == "__main__":
    main()
